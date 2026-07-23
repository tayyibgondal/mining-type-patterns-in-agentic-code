"""
Rust regex-parser evaluation harness
====================================

Evaluates the Rust *regex* type-relatedness parser (the keyword/pattern filter in
`extract_rust_type_prs.py`) against human judgement, so we can report its
accuracy / precision / recall / F1.

What it does
------------
1. Loads the FULL population of agentic Rust PRs from AIDev-pop (HuggingFace),
   i.e. every PR by an AI agent in a Rust repo (positives AND negatives).
2. Labels each PR with the regex parser's prediction:
      type_related      -> PR was flagged by the regex (present in the 790-row
                           `rust_type_prs_all_groups.csv`)
      not_type_related  -> PR was NOT flagged by the regex
   and attaches the regex evidence (detection_method, feature counts, patch).
3. Writes the full labelled dataset to
      datasets/rust_data/rust_regex_parser_predictions.csv
4. Draws a reproducible stratified sample (default 25 positive + 25 negative)
   and writes a human-evaluation workbook to
      datasets/rust_data/rust_regex_human_eval_sample.xlsx
   with independent annotation columns for two evaluators (Tayyib, Imgyeong)
   and a Metrics sheet that auto-computes accuracy/precision/recall/F1.

Evaluation model (per PR)
-------------------------
Each evaluator only decides one thing: "Is this PR genuinely type-related?" (y/n).
The confusion-matrix cell is then derived automatically:
      regex=type_related     & human=y  -> TP   (regex correctly flagged)
      regex=type_related     & human=n  -> FP   (regex over-flagged)
      regex=not_type_related & human=y  -> FN   (regex missed it)
      regex=not_type_related & human=n  -> TN   (regex correctly skipped)
"""

import os
import random

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

AI_AGENTS = ['OpenAI_Codex', 'Devin', 'Copilot', 'Cursor', 'Claude_Code']

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.normpath(os.path.join(HERE, '..', '..'))
RUST_DATA = os.path.join(REPO, 'datasets', 'rust_data')

PRED_CSV = os.path.join(RUST_DATA, 'rust_regex_parser_predictions.csv')
EVAL_XLSX = os.path.join(RUST_DATA, 'rust_regex_human_eval_sample.xlsx')

N_POS = 25   # regex-positive PRs in the sample
N_NEG = 25   # regex-negative PRs in the sample
SEED = 42

EVALUATORS = ['Tayyib', 'Imgyeong']


# ----------------------------------------------------------------------------
# 1-3. Build the full labelled predictions dataset
# ----------------------------------------------------------------------------

def build_predictions():
    print("Loading AIDev-pop tables from HuggingFace...")
    repo_df = pd.read_parquet('hf://datasets/hao-li/AIDev/repository.parquet')
    pr_df = pd.read_parquet('hf://datasets/hao-li/AIDev/pull_request.parquet')

    rust_repo_ids = set(
        repo_df[repo_df['language'].fillna('').str.contains('Rust', case=False, regex=False)]['id']
    )
    pop = pr_df[pr_df['repo_id'].isin(rust_repo_ids) & pr_df['agent'].isin(AI_AGENTS)].copy()
    print(f"Full agentic Rust PR population: {len(pop)}")

    # Regex-positive set = the PRs the regex filter flagged (all_groups file).
    allg = pd.read_csv(os.path.join(RUST_DATA, 'rust_type_prs_all_groups.csv'))
    regex_pos_ids = set(allg['id'])

    # Regex evidence + patch text for the positives.
    ev_cols = ['id', 'detection_method', 'total_feature_count',
               'unsafe_additions', 'patch_text']
    ev_cols = [c for c in ev_cols if c in allg.columns]
    evidence = allg[ev_cols].drop_duplicates('id')

    # LLM final label (for reference / comparison, NOT ground truth).
    cl = pd.read_csv(os.path.join(RUST_DATA, 'rust_classified_type_prs.csv'))
    llm = cl[['id', 'final_is_type_related']].drop_duplicates('id') \
        if 'final_is_type_related' in cl.columns else pd.DataFrame(columns=['id'])

    keep = ['id', 'number', 'title', 'body', 'agent', 'state', 'merged_at', 'html_url']
    keep = [c for c in keep if c in pop.columns]
    out = pop[keep].copy()
    out['regex_prediction'] = out['id'].apply(
        lambda i: 'type_related' if i in regex_pos_ids else 'not_type_related')
    out = out.merge(evidence, on='id', how='left')
    if not llm.empty:
        out = out.merge(llm, on='id', how='left')

    out = out.sort_values('regex_prediction').reset_index(drop=True)
    out.to_csv(PRED_CSV, index=False)
    n_pos = (out['regex_prediction'] == 'type_related').sum()
    n_neg = (out['regex_prediction'] == 'not_type_related').sum()
    print(f"Wrote {PRED_CSV}")
    print(f"  regex positive (type_related):     {n_pos}")
    print(f"  regex negative (not_type_related): {n_neg}")
    return out


# ----------------------------------------------------------------------------
# 4. Stratified sample + human-evaluation workbook
# ----------------------------------------------------------------------------

def _clip(text, n):
    s = '' if pd.isna(text) else str(text)
    s = s.replace('\r', ' ')
    return s if len(s) <= n else s[:n] + ' …[truncated]'


def sample_rows(pred):
    rng = random.Random(SEED)
    pos = pred[pred['regex_prediction'] == 'type_related']
    neg = pred[pred['regex_prediction'] == 'not_type_related']
    pos_idx = rng.sample(list(pos.index), min(N_POS, len(pos)))
    neg_idx = rng.sample(list(neg.index), min(N_NEG, len(neg)))
    sample = pred.loc[pos_idx + neg_idx].copy()
    # Shuffle so positives/negatives are interleaved (reduces ordering bias).
    order = list(sample.index)
    rng.shuffle(order)
    return pred.loc[order].reset_index(drop=True)


def build_workbook(sample):
    wb = Workbook()

    # ---- Instructions sheet ----
    ins = wb.active
    ins.title = 'Instructions'
    lines = [
        ('Rust Regex Parser — Human Evaluation', True),
        ('', False),
        ('Goal: measure how well the Rust REGEX type-relatedness parser agrees with human judgement.', False),
        ('The regex parser flags a PR as "type_related" if its Rust diff shows type-system signals', False),
        ('(traits, generics, lifetimes, unsafe, Option/Result, etc.). We check whether that is correct.', False),
        ('', False),
        ('HOW TO ANNOTATE (each evaluator, independently):', True),
        ('1. Open the PR via its html_url and read the title, body, and Rust diff.', False),
        ('2. Decide ONLY: is this PR genuinely about Rust types / the type system?  Enter y or n', False),
        ('   in your "<name>_TypeRelated" column. (Bug fixes to types, generics, traits, lifetimes,', False),
        ('   unsafe, trait bounds, enum/struct type changes = yes. Docs/CI/formatting/rename = no.)', False),
        ('3. Do NOT edit the TP/FP/TN/FN columns — they fill in automatically from your y/n answer', False),
        ('   and the regex prediction.', False),
        ('', False),
        ('Confusion matrix (auto-derived):', True),
        ('   regex=type_related     & you say y  -> TP  (regex correctly flagged a type PR)', False),
        ('   regex=type_related     & you say n  -> FP  (regex wrongly flagged a non-type PR)', False),
        ('   regex=not_type_related & you say y  -> FN  (regex missed a real type PR)', False),
        ('   regex=not_type_related & you say n  -> TN  (regex correctly skipped a non-type PR)', False),
        ('', False),
        ('The Metrics sheet computes Accuracy, Precision, Recall, F1 for each evaluator', False),
        ('once the y/n columns are filled, plus inter-rater agreement.', False),
        ('', False),
        (f'Sample: {N_POS} regex-positive + {N_NEG} regex-negative PRs (stratified, seed={SEED}),', False),
        ('drawn from the full agentic Rust PR population. Full labels: rust_regex_parser_predictions.csv', False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = ins.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=14 if (bold and i == 1) else 11)
    ins.column_dimensions['A'].width = 110

    # ---- Evaluation sheet ----
    ws = wb.create_sheet('Evaluation')
    base_cols = ['#', 'pr_id', 'number', 'agent', 'title', 'html_url',
                 'body_excerpt', 'patch_excerpt', 'regex_prediction']
    eval_cols = []
    for name in EVALUATORS:
        eval_cols += [f'{name}_TypeRelated (y/n)', f'{name}_TP', f'{name}_FP',
                      f'{name}_TN', f'{name}_FN']
    headers = base_cols + eval_cols

    header_fill = PatternFill('solid', fgColor='2C3E50')
    inp_fill = PatternFill('solid', fgColor='FFF2CC')
    pos_fill = PatternFill('solid', fgColor='FADBD8')
    neg_fill = PatternFill('solid', fgColor='D6EAF8')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    regex_col_idx = base_cols.index('regex_prediction') + 1
    regex_letter = get_column_letter(regex_col_idx)

    for i, (_, r) in enumerate(sample.iterrows(), start=2):
        vals = [
            i - 1, r['id'], r.get('number', ''), r.get('agent', ''),
            _clip(r.get('title', ''), 200), r.get('html_url', ''),
            _clip(r.get('body', ''), 600), _clip(r.get('patch_text', ''), 1500),
            r['regex_prediction'],
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=j in (5, 7, 8))
            c.border = border
        # colour-code the regex prediction cell
        ws.cell(row=i, column=regex_col_idx).fill = (
            pos_fill if r['regex_prediction'] == 'type_related' else neg_fill)

        # evaluator columns + auto formulas
        for k, name in enumerate(EVALUATORS):
            base = len(base_cols) + k * 5 + 1  # first col for this evaluator
            judg_letter = get_column_letter(base)
            ws.cell(row=i, column=base).fill = inp_fill      # y/n input
            ws.cell(row=i, column=base).border = border
            ws.cell(row=i, column=base).alignment = Alignment(horizontal='center')
            tp = (f'=IF(AND(${regex_letter}{i}="type_related",LOWER(${judg_letter}{i})="y"),1,0)')
            fp = (f'=IF(AND(${regex_letter}{i}="type_related",LOWER(${judg_letter}{i})="n"),1,0)')
            tn = (f'=IF(AND(${regex_letter}{i}="not_type_related",LOWER(${judg_letter}{i})="n"),1,0)')
            fn = (f'=IF(AND(${regex_letter}{i}="not_type_related",LOWER(${judg_letter}{i})="y"),1,0)')
            for off, formula in zip(range(1, 5), [tp, fp, tn, fn]):
                cc = ws.cell(row=i, column=base + off, value=formula)
                cc.alignment = Alignment(horizontal='center')
                cc.border = border

    # data-validation dropdowns (y/n) on the input columns
    last_row = len(sample) + 1
    for k, name in enumerate(EVALUATORS):
        base = len(base_cols) + k * 5 + 1
        letter = get_column_letter(base)
        dv = DataValidation(type='list', formula1='"y,n"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{letter}2:{letter}{last_row}')

    widths = {1: 5, 2: 12, 3: 9, 4: 14, 5: 46, 6: 42, 7: 50, 8: 60, 9: 20}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    for k in range(len(EVALUATORS)):
        base = len(base_cols) + k * 5 + 1
        ws.column_dimensions[get_column_letter(base)].width = 20
        for off in range(1, 5):
            ws.column_dimensions[get_column_letter(base + off)].width = 7
    ws.freeze_panes = 'B2'

    # ---- Metrics sheet ----
    ms = wb.create_sheet('Metrics')
    ms['A1'] = 'Regex Parser Performance (auto-computed once y/n columns are filled)'
    ms['A1'].font = Font(bold=True, size=13)
    ms.column_dimensions['A'].width = 26
    for col in 'BCDE':
        ms.column_dimensions[col].width = 16

    metric_rows = ['TP', 'FP', 'TN', 'FN', 'Total',
                   'Accuracy', 'Precision', 'Recall (Sensitivity)',
                   'Specificity', 'F1 score']
    ms.cell(row=3, column=1, value='Metric').font = Font(bold=True)
    for k, name in enumerate(EVALUATORS):
        ms.cell(row=3, column=2 + k, value=name).font = Font(bold=True)

    # references into Evaluation sheet TP/FP/TN/FN columns per evaluator
    def sums(k):
        base = len(base_cols) + k * 5 + 1
        return {t: get_column_letter(base + off)
                for t, off in zip(['TP', 'FP', 'TN', 'FN'], range(1, 5))}

    for ri, metric in enumerate(metric_rows, start=4):
        ms.cell(row=ri, column=1, value=metric).font = Font(bold=metric in
                ('Accuracy', 'Precision', 'Recall (Sensitivity)', 'F1 score'))
        for k in range(len(EVALUATORS)):
            L = sums(k)
            colL = get_column_letter(2 + k)
            tp = f"SUM(Evaluation!{L['TP']}2:{L['TP']}{last_row})"
            fp = f"SUM(Evaluation!{L['FP']}2:{L['FP']}{last_row})"
            tn = f"SUM(Evaluation!{L['TN']}2:{L['TN']}{last_row})"
            fn = f"SUM(Evaluation!{L['FN']}2:{L['FN']}{last_row})"
            if metric == 'TP':
                f = f'={tp}'
            elif metric == 'FP':
                f = f'={fp}'
            elif metric == 'TN':
                f = f'={tn}'
            elif metric == 'FN':
                f = f'={fn}'
            elif metric == 'Total':
                f = f'={tp}+{fp}+{tn}+{fn}'
            elif metric == 'Accuracy':
                f = f'=IFERROR(({tp}+{tn})/({tp}+{fp}+{tn}+{fn}),"")'
            elif metric == 'Precision':
                f = f'=IFERROR({tp}/({tp}+{fp}),"")'
            elif metric == 'Recall (Sensitivity)':
                f = f'=IFERROR({tp}/({tp}+{fn}),"")'
            elif metric == 'Specificity':
                f = f'=IFERROR({tn}/({tn}+{fp}),"")'
            else:  # F1
                f = (f'=IFERROR(2*{tp}/(2*{tp}+{fp}+{fn}),"")')
            cell = ms.cell(row=ri, column=2 + k, value=f)
            if metric in ('Accuracy', 'Precision', 'Recall (Sensitivity)',
                          'Specificity', 'F1 score'):
                cell.number_format = '0.0%'

    # inter-rater agreement (share of PRs where both y/n answers match)
    ar = len(metric_rows) + 6
    ms.cell(row=ar, column=1, value='Inter-rater agreement').font = Font(bold=True)
    j0 = get_column_letter(len(base_cols) + 1)              # Tayyib y/n
    j1 = get_column_letter(len(base_cols) + 5 + 1)          # Imgyeong y/n
    ms.cell(row=ar, column=2, value=(
        f'=IFERROR(SUMPRODUCT(--(LOWER(Evaluation!{j0}2:{j0}{last_row})='
        f'LOWER(Evaluation!{j1}2:{j1}{last_row})),'
        f'--(Evaluation!{j0}2:{j0}{last_row}<>""))'
        f'/COUNTIF(Evaluation!{j0}2:{j0}{last_row},"?*"),"")'))
    ms.cell(row=ar, column=2).number_format = '0.0%'

    wb.save(EVAL_XLSX)
    print(f"Wrote {EVAL_XLSX}  ({len(sample)} PRs: "
          f"{(sample['regex_prediction']=='type_related').sum()} positive, "
          f"{(sample['regex_prediction']=='not_type_related').sum()} negative)")


def main():
    print("=" * 70)
    print("RUST REGEX PARSER — EVALUATION HARNESS")
    print("=" * 70)
    pred = build_predictions()
    sample = sample_rows(pred)
    build_workbook(sample)
    print("\nDone. Fill the y/n columns in the Evaluation sheet; the Metrics")
    print("sheet then reports accuracy / precision / recall / F1 automatically.")


if __name__ == '__main__':
    main()
