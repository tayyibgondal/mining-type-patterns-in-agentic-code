"""
Pipeline evaluation sheets — Stage 1 (regex) and Stage 2 (LLM), all languages
============================================================================

The type-related PR pipeline has two filtering stages:

  Stage 1  REGEX parser   : scans every agentic PR in a language's repos and
                            flags those whose diff shows type-system signals.
  Stage 2  LLM classifier : re-reads the regex-flagged PRs and decides which
                            are *genuinely* type-related (`final_is_type_related`).

This script builds a human-evaluation workbook for each (language, stage) so we
can measure accuracy / precision / recall / F1 of each stage independently.

Populations
-----------
Stage 1 (regex): the FULL agentic PR population for the language, taken from
    AIDev-pop on HuggingFace. Positive = PR was flagged by the regex (present in
    the regex-output CSV); negative = it was not.
Stage 2 (LLM): the regex-flagged set only (that is the LLM's actual input).
    Positive = `final_is_type_related == True`; negative = False.
    NOTE: stage-2 metrics are therefore *conditional* on the regex having
    flagged the PR — that is the correct denominator for judging stage 2.

Each workbook has a stratified 25 positive + 25 negative sample, independent
y/n columns for two evaluators, auto-derived TP/FP/TN/FN, and a Metrics sheet.
"""

import os
import random

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

AI_AGENTS = ['OpenAI_Codex', 'Devin', 'Copilot', 'Cursor', 'Claude_Code']

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.normpath(os.path.join(HERE, '..'))
DATASETS = os.path.join(REPO, 'datasets')

N_POS = 25
N_NEG = 25
SEED = 42
EVALUATORS = ['Tayyib', 'Imgyeong']

LANGS = {
    'rust': dict(dir='rust_data', aidev='Rust', prefix='rust',
                 regex_pos='rust_type_prs_all_groups.csv',
                 classified='rust_classified_type_prs.csv',
                 src_ext='.rs', display='Rust',
                 signals='traits, generics, lifetimes, unsafe, Option/Result'),
    'typescript': dict(dir='typescript_data', aidev='TypeScript', prefix='typescript',
                       regex_pos='agent_type_prs_unfiltered.csv',
                       classified='agent_type_prs_filtered_by_open_ai.csv',
                       src_ext='.ts', display='TypeScript',
                       signals='`any`, generics, union/utility types, type guards, assertions'),
    'csharp': dict(dir='csharp_data', aidev='C#', prefix='csharp',
                   regex_pos='csharp_type_prs_all_groups.csv',
                   classified='csharp_classified_type_prs.csv',
                   src_ext='.cs', display='C#',
                   signals='`dynamic`, generics, nullable types, pattern matching, records'),
}

_AIDEV = {}


def aidev():
    """Load and cache the AIDev-pop repository/pull_request tables."""
    if not _AIDEV:
        print("  loading AIDev-pop from HuggingFace ...")
        _AIDEV['repo'] = pd.read_parquet('hf://datasets/hao-li/AIDev/repository.parquet')
        _AIDEV['pr'] = pd.read_parquet('hf://datasets/hao-li/AIDev/pull_request.parquet')
    return _AIDEV['repo'], _AIDEV['pr']


def _clip(text, n):
    s = '' if pd.isna(text) else str(text)
    s = s.replace('\r', ' ')
    return s if len(s) <= n else s[:n] + ' …[truncated]'


# ---------------------------------------------------------------------------
# Population builders
# ---------------------------------------------------------------------------

def build_regex_population(cfg):
    repo_df, pr_df = aidev()
    repo_ids = set(repo_df[repo_df['language'].fillna('') == cfg['aidev']]['id'])
    pop = pr_df[pr_df['repo_id'].isin(repo_ids) & pr_df['agent'].isin(AI_AGENTS)].copy()

    pos = pd.read_csv(os.path.join(DATASETS, cfg['dir'], cfg['regex_pos']), low_memory=False)
    pos_ids = set(pos['id'])

    ev_cols = [c for c in ['id', 'detection_method', 'total_feature_count', 'patch_text']
               if c in pos.columns]
    evidence = pos[ev_cols].drop_duplicates('id')

    keep = [c for c in ['id', 'number', 'title', 'body', 'agent', 'state',
                        'merged_at', 'html_url'] if c in pop.columns]
    out = pop[keep].copy()
    out['prediction'] = out['id'].apply(
        lambda i: 'type_related' if i in pos_ids else 'not_type_related')
    out = out.merge(evidence, on='id', how='left')
    return out


def build_llm_population(cfg):
    cl = pd.read_csv(os.path.join(DATASETS, cfg['dir'], cfg['classified']), low_memory=False)
    keep = [c for c in ['id', 'number', 'title', 'body', 'agent', 'state', 'merged_at',
                        'html_url', 'patch_text', 'classifier_reasoning',
                        'final_is_type_related'] if c in cl.columns]
    out = cl[keep].copy()
    out['prediction'] = out['final_is_type_related'].apply(
        lambda v: 'type_related' if v is True or str(v).lower() == 'true' else 'not_type_related')
    return out


def sample_rows(pop):
    rng = random.Random(SEED)
    pos = pop[pop['prediction'] == 'type_related']
    neg = pop[pop['prediction'] == 'not_type_related']
    idx = (rng.sample(list(pos.index), min(N_POS, len(pos))) +
           rng.sample(list(neg.index), min(N_NEG, len(neg))))
    rng.shuffle(idx)
    return pop.loc[idx].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def build_workbook(sample, cfg, stage, out_path):
    is_regex = stage == 'regex'
    stage_name = 'REGEX parser (stage 1)' if is_regex else 'LLM classifier (stage 2)'
    pred_header = 'regex_prediction' if is_regex else 'llm_prediction'
    lang = cfg['display']

    wb = Workbook()
    ins = wb.active
    ins.title = 'Instructions'
    if is_regex:
        scope = [
            (f'Scope: every AI-agent PR in {lang} repositories (AIDev-pop). The regex parser', False),
            (f'flags a PR as "type_related" when its {cfg["src_ext"]} diff shows type signals', False),
            (f'({cfg["signals"]}).', False),
        ]
    else:
        scope = [
            (f'Scope: only PRs the REGEX already flagged — that is the LLM\'s actual input.', False),
            ('The LLM classifier re-read each one and decided whether it is genuinely', False),
            ('type-related. Metrics here are therefore conditional on stage 1 having flagged it.', False),
        ]
    lines = [
        (f'{lang} — {stage_name} — Human Evaluation', True),
        ('', False),
        (f'Goal: measure how well the {stage_name} agrees with human judgement.', False),
    ] + scope + [
        ('', False),
        ('HOW TO ANNOTATE (each evaluator, independently):', True),
        ('1. Open the PR via its html_url (and read the patch_excerpt) — title, body, diff.', False),
        (f'2. Decide ONLY: is this PR genuinely about {lang} types / the type system?', False),
        ('   Enter y or n in your "<name>_TypeRelated" column.', False),
        ('   YES = type/generic/trait/interface definitions, type annotations & signatures,', False),
        ('         type-safety fixes, escape hatches (any / dynamic / unsafe), type conversions.', False),
        ('   NO  = feature/logic/perf work, docs, comments, tests, CI, formatting, renames,', False),
        ('         or code where types are only incidental plumbing.', False),
        ('3. Do NOT edit the TP/FP/TN/FN columns — they fill in automatically.', False),
        ('', False),
        ('Confusion matrix (auto-derived):', True),
        (f'   {pred_header}=type_related     & you say y  -> TP', False),
        (f'   {pred_header}=type_related     & you say n  -> FP', False),
        (f'   {pred_header}=not_type_related & you say y  -> FN', False),
        (f'   {pred_header}=not_type_related & you say n  -> TN', False),
        ('', False),
        ('The Metrics sheet computes Accuracy, Precision, Recall, Specificity, F1 per', False),
        ('evaluator plus inter-rater agreement, once the y/n columns are filled.', False),
        ('', False),
        (f'Sample: {N_POS} predicted-positive + {N_NEG} predicted-negative (stratified, seed={SEED}).', False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        c = ins.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=14 if (bold and i == 1) else 11)
    ins.column_dimensions['A'].width = 112

    ws = wb.create_sheet('Evaluation')
    base_cols = ['#', 'pr_id', 'number', 'agent', 'title', 'html_url',
                 'body_excerpt', 'patch_excerpt', pred_header]
    eval_cols = []
    for name in EVALUATORS:
        eval_cols += [f'{name}_TypeRelated (y/n)', f'{name}_TP', f'{name}_FP',
                      f'{name}_TN', f'{name}_FN']
    headers = base_cols + eval_cols

    header_fill = PatternFill('solid', fgColor='2C3E50')
    inp_fill = PatternFill('solid', fgColor='FFF2CC')
    pos_fill = PatternFill('solid', fgColor='FADBD8')
    neg_fill = PatternFill('solid', fgColor='D6EAF8')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    pred_idx = base_cols.index(pred_header) + 1
    pred_letter = get_column_letter(pred_idx)

    for i, (_, r) in enumerate(sample.iterrows(), start=2):
        vals = [i - 1, r['id'], r.get('number', ''), r.get('agent', ''),
                _clip(r.get('title', ''), 200), r.get('html_url', ''),
                _clip(r.get('body', ''), 600), _clip(r.get('patch_text', ''), 1500),
                r['prediction']]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=j in (5, 7, 8))
            c.border = border
        ws.cell(row=i, column=pred_idx).fill = (
            pos_fill if r['prediction'] == 'type_related' else neg_fill)

        for k in range(len(EVALUATORS)):
            base = len(base_cols) + k * 5 + 1
            jl = get_column_letter(base)
            cell = ws.cell(row=i, column=base)
            cell.fill = inp_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            fs = [
                f'=IF(AND(${pred_letter}{i}="type_related",LOWER(${jl}{i})="y"),1,0)',
                f'=IF(AND(${pred_letter}{i}="type_related",LOWER(${jl}{i})="n"),1,0)',
                f'=IF(AND(${pred_letter}{i}="not_type_related",LOWER(${jl}{i})="n"),1,0)',
                f'=IF(AND(${pred_letter}{i}="not_type_related",LOWER(${jl}{i})="y"),1,0)',
            ]
            for off, formula in zip(range(1, 5), fs):
                cc = ws.cell(row=i, column=base + off, value=formula)
                cc.alignment = Alignment(horizontal='center')
                cc.border = border

    last_row = len(sample) + 1
    for k in range(len(EVALUATORS)):
        base = len(base_cols) + k * 5 + 1
        letter = get_column_letter(base)
        dv = DataValidation(type='list', formula1='"y,n"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{letter}2:{letter}{last_row}')

    for idx, w in {1: 5, 2: 12, 3: 9, 4: 14, 5: 46, 6: 42, 7: 50, 8: 60, 9: 20}.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    for k in range(len(EVALUATORS)):
        base = len(base_cols) + k * 5 + 1
        ws.column_dimensions[get_column_letter(base)].width = 20
        for off in range(1, 5):
            ws.column_dimensions[get_column_letter(base + off)].width = 7
    ws.freeze_panes = 'B2'

    ms = wb.create_sheet('Metrics')
    ms['A1'] = f'{lang} — {stage_name} performance (auto-computed)'
    ms['A1'].font = Font(bold=True, size=13)
    ms.column_dimensions['A'].width = 26
    for col in 'BCDE':
        ms.column_dimensions[col].width = 16

    metric_rows = ['TP', 'FP', 'TN', 'FN', 'Total', 'Accuracy', 'Precision',
                   'Recall (Sensitivity)', 'Specificity', 'F1 score']
    ms.cell(row=3, column=1, value='Metric').font = Font(bold=True)
    for k, name in enumerate(EVALUATORS):
        ms.cell(row=3, column=2 + k, value=name).font = Font(bold=True)

    def L(k):
        base = len(base_cols) + k * 5 + 1
        return {t: get_column_letter(base + off)
                for t, off in zip(['TP', 'FP', 'TN', 'FN'], range(1, 5))}

    for ri, metric in enumerate(metric_rows, start=4):
        ms.cell(row=ri, column=1, value=metric).font = Font(
            bold=metric in ('Accuracy', 'Precision', 'Recall (Sensitivity)', 'F1 score'))
        for k in range(len(EVALUATORS)):
            c = L(k)
            tp = f"SUM(Evaluation!{c['TP']}2:{c['TP']}{last_row})"
            fp = f"SUM(Evaluation!{c['FP']}2:{c['FP']}{last_row})"
            tn = f"SUM(Evaluation!{c['TN']}2:{c['TN']}{last_row})"
            fn = f"SUM(Evaluation!{c['FN']}2:{c['FN']}{last_row})"
            f = {
                'TP': f'={tp}', 'FP': f'={fp}', 'TN': f'={tn}', 'FN': f'={fn}',
                'Total': f'={tp}+{fp}+{tn}+{fn}',
                'Accuracy': f'=IFERROR(({tp}+{tn})/({tp}+{fp}+{tn}+{fn}),"")',
                'Precision': f'=IFERROR({tp}/({tp}+{fp}),"")',
                'Recall (Sensitivity)': f'=IFERROR({tp}/({tp}+{fn}),"")',
                'Specificity': f'=IFERROR({tn}/({tn}+{fp}),"")',
                'F1 score': f'=IFERROR(2*{tp}/(2*{tp}+{fp}+{fn}),"")',
            }[metric]
            cell = ms.cell(row=ri, column=2 + k, value=f)
            if metric in ('Accuracy', 'Precision', 'Recall (Sensitivity)',
                          'Specificity', 'F1 score'):
                cell.number_format = '0.0%'

    ar = len(metric_rows) + 6
    ms.cell(row=ar, column=1, value='Inter-rater agreement').font = Font(bold=True)
    j0 = get_column_letter(len(base_cols) + 1)
    j1 = get_column_letter(len(base_cols) + 5 + 1)
    ms.cell(row=ar, column=2, value=(
        f'=IFERROR(SUMPRODUCT(--(LOWER(Evaluation!{j0}2:{j0}{last_row})='
        f'LOWER(Evaluation!{j1}2:{j1}{last_row})),'
        f'--(Evaluation!{j0}2:{j0}{last_row}<>""))'
        f'/COUNTIF(Evaluation!{j0}2:{j0}{last_row},"?*"),"")'))
    ms.cell(row=ar, column=2).number_format = '0.0%'

    wb.save(out_path)
    return out_path


def generate(lang_key, stage):
    cfg = LANGS[lang_key]
    d = os.path.join(DATASETS, cfg['dir'])
    print(f"\n[{cfg['display']} / {stage}]")
    pop = build_regex_population(cfg) if stage == 'regex' else build_llm_population(cfg)
    n_pos = (pop['prediction'] == 'type_related').sum()
    n_neg = (pop['prediction'] == 'not_type_related').sum()
    print(f"  population={len(pop)}  positive={n_pos}  negative={n_neg}")

    csv_path = os.path.join(d, f"{cfg['prefix']}_{stage}_predictions.csv")
    cols = [c for c in pop.columns if c != 'patch_text']  # keep the CSV lean
    pop[cols].to_csv(csv_path, index=False)

    sample = sample_rows(pop)
    xlsx = os.path.join(d, f"{cfg['prefix']}_{stage}_human_eval_sample.xlsx")
    build_workbook(sample, cfg, stage, xlsx)
    print(f"  wrote {os.path.basename(csv_path)} and {os.path.basename(xlsx)} "
          f"({len(sample)} sampled)")
    return xlsx


if __name__ == '__main__':
    import sys
    # Rust/regex is intentionally excluded by default: that workbook already
    # exists and contains completed annotations.
    todo = [('typescript', 'regex'), ('csharp', 'regex'),
            ('rust', 'llm'), ('typescript', 'llm'), ('csharp', 'llm')]
    if len(sys.argv) > 1:
        todo = [tuple(a.split(':')) for a in sys.argv[1:]]
    for lang, stage in todo:
        generate(lang, stage)
    print("\nDone.")
